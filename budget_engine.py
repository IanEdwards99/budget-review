from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BUDGET = {
    "Rent": 1537.80,
    "Groceries/Household": 550.00,
    "Internet (Odido)": 34.00,
    "Cell Phone": 46.02,
    "Transport": 80.00,
    "Electricity/Gas": 110.00,
    "Water": 15.00,
    "Gym": 0.00,
    "Ian - Fun": 250.00,
    "Leila - Fun": 250.00,
    "Investments": 400.00,
    "Other Shopping": 0.00,
    "Bank Fees": 10.00,
    "Unexpected/Emergencies": 100.00,
}

INCOME_BUDGET = {
    "Monthly Salary (Ian)": 3500.00,
    "Monthly Salary (Leila)": 1550.00,
}

UTILITY_CATEGORIES = ["Salary", "Other Income / Refunds", "Internal Transfer"]
ALL_CATEGORIES = list(BUDGET) + UTILITY_CATEGORIES

COLORS = {
    "dark": "1F3864",
    "mid": "2E75B6",
    "light": "DDEBF7",
    "stripe": "EBF3FB",
    "grey": "F2F2F2",
    "green": "E2F0D9",
    "red": "FCE4D6",
    "border": "A6A6A6",
}


@dataclass
class Period:
    name: str
    start: date
    end: date
    data_through: date

    @property
    def elapsed_pct(self) -> float:
        return round(((self.data_through - self.start).days + 1) / ((self.end - self.start).days + 1), 4)


def clean_spaces(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def safe_sheet_name(text: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "-", text)[:31]


def abn_merchant(description: str) -> str:
    match = re.search(r"/NAME/([^/]+)", description)
    if match:
        return clean_spaces(match.group(1))
    match = re.search(r"(?:BEA|eCom),\s*[^ ]+\s+(.+?)(?:,PAS|\s+NR:)", description)
    if match:
        merchant = clean_spaces(match.group(1))
        merchant = re.sub(r"^(?:Pay\s+)+", "", merchant, flags=re.I).strip()
        merchant = re.sub(r"^(?:BCK\*|CCV\*|PAY\*|PAY\.nl\*)", "", merchant, flags=re.I).strip()
        return merchant
    return clean_spaces(description[:80])


def category_from_text(text: str, source_category: str = "") -> tuple[str, bool, str]:
    t = f"{text} {source_category}".lower()

    if any(x in t for x in ["asml netherlands", "salaris", "salary/payroll"]):
        return "Salary", True, "Salary detected from statement"

    looks_like_ian_counterparty = (
        t.strip() in {"ian edwards", "ian michael edwards"}
        or t.startswith("ian edwards ")
        or t.startswith("ian michael edwards ")
        or "/name/ian edwards" in t
        or "/name/ian michael edwards" in t
        or "wise transfer" in t
    )
    if looks_like_ian_counterparty or any(x in t for x in ["money added", "pay back fun money"]):
        return "Internal Transfer", False, "Excluded to avoid double-counting account movement"

    if "leila fun money" in t or "leila siljeur" in t:
        return "Leila - Fun", True, "Leila fun-money allocation from joint account"

    rules = [
        ("Rent", ["project 434", "nieuwe fellenoord"]),
        ("Internet (Odido)", ["odido internet", "odido int", "yll"]),
        ("Cell Phone", ["odido mobile", "mob 06", "mobile"]),
        ("Electricity/Gas", ["energiedirect"]),
        ("Water", ["brabant water"]),
        ("Bank Fees", ["basispakket", "paypal", "transferwise", "assets fee", "wise assets fee", "abn amro bank"]),
        ("Investments", ["etoro", "investment", "investments"]),
        ("Gym", ["boulder", "neoliet"]),
        (
            "Groceries/Household",
            ["albert heijn", "ah to go", "hizmet", "diyar", "sang lee", "dekamarkt", "supermarkt", "lidl", "kruidvat", "normal eindhove", "die spens"],
        ),
        (
            "Transport",
            ["ns groep", "ns reizigers", "transport", "flix", "uber", "esso", "total service station", "parkeergarage", "eventbike", "city bike", "db reisezentrum", "den haag cs"],
        ),
        (
            "Other Shopping",
            ["amazon", "hm online", "h.m online", "h&m", "wibra", "kik", "primark", "decathlon", "pipoos", "magicers", "openai", "chat gpt", "claude", "anthropic", "all4running", "shopping"],
        ),
        ("Unexpected/Emergencies", ["zilveren kruis", "belastingdienst", "gemeente", "ambassade", "robert", "dad", "family", "healthcare"]),
    ]
    for category, terms in rules:
        if any(term in t for term in terms):
            return category, True, ""

    if source_category in {"Bills"}:
        return "Bank Fees", True, "Wise category fallback"
    if source_category in {"Groceries"}:
        return "Groceries/Household", True, "Wise category fallback"
    if source_category in {"Transport"}:
        return "Transport", True, "Wise category fallback"
    if source_category in {"Investments"}:
        return "Investments", True, "Wise category fallback"
    if source_category in {"Shopping", "Personal care"}:
        return "Other Shopping", True, "Wise category fallback"
    if source_category in {"Eating out", "Entertainment", "Trips", "Cash", "General"}:
        return "Ian - Fun", True, "Wise category fallback"

    fun_terms = ["booking", "hotel", "trip", "eating out", "entertainment", "starbucks", "coffee", "cinema", "vue", "restaurant", "mcdonald", "burger", "cash", "tikkie", "vriendenloterij", "kiosk", "june tea", "ijssalon", "jolie spellen"]
    if any(term in t for term in fun_terms):
        return "Ian - Fun", True, ""

    return "Ian - Fun", True, "Default review category; check if material"


def assign_period(tx_date: date, period: Period, category: str) -> str | None:
    if period.start <= tx_date <= period.end:
        return period.name
    if category == "Salary" and 0 <= (period.start - tx_date).days <= 4:
        return period.name
    return None


def make_row(tx_date: date, period: Period, source: str, direction: str, merchant: str, description: str, signed_amount: float, category: str, include: bool, notes: str, source_category: str = "", source_id: str = "") -> dict[str, Any] | None:
    period_name = assign_period(tx_date, period, category)
    if not period_name:
        return None
    signed_amount = round(float(signed_amount), 2)
    return {
        "date": tx_date.isoformat(),
        "period": period_name,
        "source": source,
        "direction": direction,
        "merchant": clean_spaces(merchant),
        "description": clean_spaces(description),
        "amount_signed": signed_amount,
        "expense": round(abs(signed_amount), 2) if signed_amount < 0 and include else 0.0,
        "credit": round(signed_amount, 2) if signed_amount > 0 and include else 0.0,
        "category": category,
        "include": 1 if include else 0,
        "notes": notes,
        "source_category": source_category or "",
        "source_id": source_id or "",
    }


def read_abn(path: Path, period: Period) -> list[dict[str, Any]]:
    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    df = pd.read_excel(path, sheet_name=0, engine=engine)
    rows: list[dict[str, Any]] = []
    for idx, item in df.iterrows():
        if "transactiondate" not in df.columns or "amount" not in df.columns or "description" not in df.columns:
            raise ValueError("ABN file must contain transactiondate, amount, and description columns.")
        tx_date = datetime.strptime(str(int(item["transactiondate"])), "%Y%m%d").date()
        description = str(item["description"])
        merchant = abn_merchant(description)
        signed = float(item["amount"])
        direction = "IN" if signed > 0 else "OUT"
        if signed > 0:
            category, include, notes = category_from_text(description)
            if category not in {"Salary", "Internal Transfer"}:
                category, include, notes = "Other Income / Refunds", True, "Credit/refund excluded from expenses"
        else:
            category, include, notes = category_from_text(f"{merchant} {description}")
        row = make_row(tx_date, period, "ABN AMRO", direction, merchant, description, signed, category, include, notes, "", f"ABN-{idx}")
        if row:
            rows.append(row)
    return rows


def read_wise(path: Path, period: Period) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path)
    elif suffix == ".xls":
        df = pd.read_excel(path, sheet_name=0, engine="xlrd")
    elif suffix == ".xlsx":
        df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    else:
        raise ValueError("Wise file must be .csv, .xls, or .xlsx.")
    required = {"Finished on", "Created on", "Direction", "Source amount (after fees)", "Target name", "Category"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Wise file is missing required columns: {', '.join(missing)}")
    df["date"] = pd.to_datetime(df["Finished on"].fillna(df["Created on"]))
    rows: list[dict[str, Any]] = []
    for idx, item in df.iterrows():
        tx_date = item["date"].date()
        direction = str(item["Direction"])
        source_category = "" if pd.isna(item.get("Category")) else str(item.get("Category"))
        merchant = "" if pd.isna(item.get("Target name")) else str(item.get("Target name"))
        reference = "" if pd.isna(item.get("Reference")) else str(item.get("Reference"))
        description = " | ".join(x for x in [merchant, reference] if x)
        amount = float(item["Source amount (after fees)"] or 0)
        if abs(amount) < 0.005:
            continue
        if direction == "OUT":
            signed = -amount
            category, include, notes = category_from_text(f"{merchant} {reference}", source_category)
            if category == "Ian - Fun":
                notes = notes or "Wise fun spend assigned to Ian - Fun"
        elif direction == "IN":
            signed = amount
            category, include, notes = category_from_text(f"{merchant} {reference}", source_category)
            if category not in {"Salary", "Internal Transfer"}:
                category, include, notes = "Other Income / Refunds", True, "Wise incoming credit/refund"
        else:
            signed, category, include, notes = 0.0, "Internal Transfer", False, "Wise neutral movement"
        row = make_row(tx_date, period, "Wise", direction, merchant, description, signed, category, include, notes, source_category, str(item.get("ID", idx)))
        if row:
            rows.append(row)
    return rows


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def thin_border(color: str = "D9E2EC") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_header(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill(COLORS["mid"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border(COLORS["border"])


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def table(ws, ref: str, name: str) -> None:
    tab = Table(displayName=re.sub(r"[^A-Za-z0-9_]", "", name)[:254], ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def replace_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def build_workbook(rows: list[dict[str, Any]], period: Period, output_path: Path, template_path: Path | None = None) -> dict[str, Any]:
    wb = load_workbook(template_path) if template_path and template_path.exists() else Workbook()
    if wb.sheetnames == ["Sheet"]:
        del wb["Sheet"]

    settings = replace_sheet(wb, "Settings 2026")
    raw_ws = replace_sheet(wb, "Raw Txns 2026")
    review_name = safe_sheet_name(f"Review {period.name}")
    review = replace_sheet(wb, review_name)

    for ws in [settings, raw_ws, review]:
        ws.sheet_view.showGridLines = False

    set_widths(settings, {"A": 28, "B": 16, "C": 22, "D": 72, "E": 12})
    settings.merge_cells("A1:E1")
    settings["A1"] = "Budget Review Settings 2026"
    settings["A1"].fill = fill(COLORS["dark"])
    settings["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    settings["A1"].alignment = Alignment(horizontal="center")
    settings.append([])
    settings.append([])
    settings.append(["Category", "Monthly Budget", "Type", "Notes", "Active"])
    style_header(settings, 4, 1, 5)
    for cat in ALL_CATEGORIES:
        settings.append([cat, BUDGET.get(cat), "Income/Internal" if cat in UTILITY_CATEGORIES else "Expense", "", True])
    for row in settings.iter_rows(min_row=5, max_row=4 + len(ALL_CATEGORIES), min_col=1, max_col=5):
        for c in row:
            c.border = thin_border()
        row[1].number_format = "#,##0.00"
    leila_row = 8 + len(ALL_CATEGORIES)
    settings.cell(leila_row, 1, "Income Assumption")
    settings.cell(leila_row, 2, "Amount")
    style_header(settings, leila_row, 1, 2)
    settings.cell(leila_row + 1, 1, "Monthly Salary (Ian) fallback")
    settings.cell(leila_row + 1, 2, INCOME_BUDGET["Monthly Salary (Ian)"])
    settings.cell(leila_row + 2, 1, "Monthly Salary (Leila) assumption")
    settings.cell(leila_row + 2, 2, INCOME_BUDGET["Monthly Salary (Leila)"])
    settings.freeze_panes = "A5"

    set_widths(raw_ws, {"A": 12, "B": 15, "C": 12, "D": 10, "E": 28, "F": 70, "G": 14, "H": 14, "I": 14, "J": 24, "K": 10, "L": 42, "M": 18, "N": 22})
    raw_headers = ["Date", "Period", "Source", "Direction", "Merchant", "Description", "Signed Amount", "Expense", "Credit", "Category", "Include", "Notes", "Source Category", "Source ID"]
    raw_ws.append(raw_headers)
    style_header(raw_ws, 1, 1, len(raw_headers))
    for item in sorted(rows, key=lambda r: (r["date"], r["source"], r["source_id"])):
        raw_ws.append([
            item["date"], item["period"], item["source"], item["direction"], item["merchant"], item["description"],
            item["amount_signed"], item["expense"], item["credit"], item["category"], item["include"], item["notes"], item["source_category"], item["source_id"],
        ])
    for row in raw_ws.iter_rows(min_row=2, max_row=raw_ws.max_row, min_col=1, max_col=len(raw_headers)):
        for c in row:
            c.border = thin_border()
        for c in row[6:9]:
            c.number_format = "#,##0.00;(#,##0.00);-"
    if raw_ws.max_row > 1:
        table(raw_ws, f"A1:N{raw_ws.max_row}", "RawTransactions2026")
    raw_ws.freeze_panes = "A2"

    set_widths(review, {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14, "F": 18, "G": 18, "H": 58})
    review.merge_cells("A1:H1")
    review["A1"] = f"Household Budget Review | {period.name}"
    review["A1"].fill = fill(COLORS["dark"])
    review["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    review["A1"].alignment = Alignment(horizontal="center")
    review.merge_cells("A2:H2")
    review["A2"] = f"Budget cycle {period.start.isoformat()} to {period.end.isoformat()}; data through {period.data_through.isoformat()}."
    review["A2"].font = Font(italic=True, color="666666")
    review["A2"].alignment = Alignment(horizontal="center")
    review["A3"], review["B3"], review["D3"], review["E3"] = "Period key", period.name, "Elapsed", period.elapsed_pct
    review["E3"].number_format = "0%"
    category_start_row = 12
    category_end_row = category_start_row + len(BUDGET) - 1
    total_row = category_end_row + 1

    summary_rows = [
        ("Ian salary actual", "=SUMIFS('Raw Txns 2026'!$I:$I,'Raw Txns 2026'!$B:$B,$B$3,'Raw Txns 2026'!$J:$J,\"Salary\",'Raw Txns 2026'!$K:$K,1)", "Monthly expense budget", f"=SUM(B{category_start_row}:B{category_end_row})"),
        ("Leila salary assumption", f"='Settings 2026'!B{leila_row + 2}", "Actual expenses", f"=SUM(C{category_start_row}:C{category_end_row})"),
        ("Other credits/refunds", "=SUMIFS('Raw Txns 2026'!$I:$I,'Raw Txns 2026'!$B:$B,$B$3,'Raw Txns 2026'!$J:$J,\"Other Income / Refunds\",'Raw Txns 2026'!$K:$K,1)", "Expense variance", "=E4-E5"),
        ("Total income/credits", "=SUM(B4:B6)", "Budget net", "=B7-E4"),
        ("Actual net after expenses", "=B7-E5", "Net vs budget", "=B8-E7"),
    ]
    for offset, row in enumerate(summary_rows, 4):
        review.cell(offset, 1, row[0])
        review.cell(offset, 2, row[1])
        review.cell(offset, 4, row[2])
        review.cell(offset, 5, row[3])
        for col in [1, 4]:
            review.cell(offset, col).fill = fill(COLORS["light"])
            review.cell(offset, col).font = Font(bold=True)
        for col in [2, 5]:
            review.cell(offset, col).number_format = "#,##0.00;(#,##0.00);-"

    header_row = 11
    review.append([])
    review.cell(header_row, 1, "Category")
    for i, heading in enumerate(["Budget", "Actual", "Variance", "Status", "Pace"], 2):
        review.cell(header_row, i, heading)
    style_header(review, header_row, 1, 6)
    for idx, cat in enumerate(BUDGET, category_start_row):
        settings_row = 5 + list(BUDGET).index(cat)
        review.cell(idx, 1, cat)
        review.cell(idx, 2, f"='Settings 2026'!B{settings_row}")
        review.cell(idx, 3, f"=SUMIFS('Raw Txns 2026'!$H:$H,'Raw Txns 2026'!$B:$B,$B$3,'Raw Txns 2026'!$J:$J,\"{cat}\",'Raw Txns 2026'!$K:$K,1)")
        review.cell(idx, 4, f"=B{idx}-C{idx}")
        review.cell(idx, 5, f'=IF(B{idx}=0,IF(C{idx}=0,"OK","Unbudgeted"),IF(C{idx}<=B{idx},"Under","Over"))')
        review.cell(idx, 6, f'=IF($E$3>=1,E{idx},IF(B{idx}=0,E{idx},IF(C{idx}<=B{idx}*$E$3,"On pace","Ahead")))')
        for col in range(1, 7):
            review.cell(idx, col).border = thin_border()
            if idx % 2 == 0:
                review.cell(idx, col).fill = fill(COLORS["stripe"])
        for col in [2, 3, 4]:
            review.cell(idx, col).number_format = "#,##0.00;(#,##0.00);-"
    review.cell(total_row, 1, "TOTAL")
    review.cell(total_row, 2, f"=SUM(B{category_start_row}:B{category_end_row})")
    review.cell(total_row, 3, f"=SUM(C{category_start_row}:C{category_end_row})")
    review.cell(total_row, 4, f"=B{total_row}-C{total_row}")
    for col in range(1, 7):
        review.cell(total_row, col).fill = fill(COLORS["light"])
        review.cell(total_row, col).font = Font(bold=True)
        review.cell(total_row, col).border = thin_border(COLORS["border"])

    detail_start = 29
    review.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=8)
    review.cell(detail_start, 1, "Transactions included in expense totals")
    review.cell(detail_start, 1).fill = fill(COLORS["dark"])
    review.cell(detail_start, 1).font = Font(bold=True, color="FFFFFF")
    detail_rows = [r for r in rows if r["period"] == period.name and r["include"] == 1 and r["expense"] > 0]
    detail_headers = ["Date", "Source", "Merchant", "Expense", "Direction", "Notes", "Description", "Source ID"]
    row_idx = detail_start + 1
    for category in BUDGET:
        category_rows = sorted(
            [r for r in detail_rows if r["category"] == category],
            key=lambda r: (r["date"], -r["expense"], r["merchant"]),
        )
        if not category_rows:
            continue

        review.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
        review.cell(row_idx, 1, f"{category} | {len(category_rows)} transactions | EUR {sum(r['expense'] for r in category_rows):,.2f}")
        review.cell(row_idx, 1).fill = fill(COLORS["light"])
        review.cell(row_idx, 1).font = Font(bold=True, color="1F3864")
        review.cell(row_idx, 1).border = thin_border(COLORS["border"])
        row_idx += 1

        for col, heading in enumerate(detail_headers, 1):
            review.cell(row_idx, col, heading)
        style_header(review, row_idx, 1, 8)
        row_idx += 1

        first_txn_row = row_idx
        for item in category_rows:
            values = [item["date"], item["source"], item["merchant"], item["expense"], item["direction"], item["notes"], item["description"][:120], item["source_id"]]
            for col, value in enumerate(values, 1):
                review.cell(row_idx, col, value)
                review.cell(row_idx, col).border = thin_border()
                if row_idx % 2 == 0:
                    review.cell(row_idx, col).fill = fill(COLORS["stripe"])
            review.cell(row_idx, 4).number_format = "#,##0.00"
            row_idx += 1

        review.cell(row_idx, 1, "Subtotal")
        review.cell(row_idx, 4, f"=SUM(D{first_txn_row}:D{row_idx - 1})")
        for col in range(1, 9):
            review.cell(row_idx, col).fill = fill(COLORS["grey"])
            review.cell(row_idx, col).font = Font(bold=True)
            review.cell(row_idx, col).border = thin_border(COLORS["border"])
        review.cell(row_idx, 4).number_format = "#,##0.00"
        row_idx += 2
    review.freeze_panes = "A12"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    included = pd.DataFrame(rows)
    expense_total = float(included.query("include == 1")["expense"].sum()) if len(included) else 0.0
    credits_total = float(included.query("include == 1")["credit"].sum()) if len(included) else 0.0
    return {"workbook": str(output_path), "rows": len(rows), "expense_total": round(expense_total, 2), "credits_total": round(credits_total, 2), "review_sheet": review_name}


def workbook_payload_for_apps_script(workbook_path: Path, sheet_names: list[str]) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=False)
    sheets = []
    for name in sheet_names:
        ws = wb[name]
        matrix = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = []
            for cell in row:
                value = cell.value
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                values.append(value)
            matrix.append(values)
        sheets.append({"name": name, "values": matrix})
    return sheets


def save_summary_json(path: Path, result: dict[str, Any]) -> None:
    path.write_text(json.dumps(result, indent=2), encoding="utf-8")
